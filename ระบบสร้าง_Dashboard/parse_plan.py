# -*- coding: utf-8 -*-
"""
parse_plan.py
อ่านโครงสร้างแผนปฏิบัติการบริหารความเสี่ยง จาก Sheet 'Risk_Bar'
แล้วสร้าง plan_data.json  (โครงสร้างแผน + แผน % รายเดือน baseline)
ระดับชั้น:  section -> subplan -> measure -> task
"""
import openpyxl, json, re, sys

import os
HERE=os.path.dirname(os.path.abspath(__file__))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE,"แผนปฏิบัติการบริหารความเสี่ยง_2569.xlsx")
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Risk_Bar"]

MONTHS = ["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.","ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]
G, R = 7, 18  # monthly cols (G..R)

def cell(r, c):
    return ws.cell(row=r, column=c).value

def monthly(r):
    """คืน list 12 ค่า (float หรือ None) จากคอลัมน์ G..R"""
    out = []
    for c in range(G, R+1):
        v = cell(r, c)
        try:
            out.append(round(float(v), 1) if v is not None else None)
        except (TypeError, ValueError):
            out.append(None)
    return out

def is_pct_row(r):
    """แถวที่เป็นเปอร์เซ็นต์สะสม (มีค่าทศนิยม/มากกว่า 12)"""
    vals = [cell(r, c) for c in range(G, R+1)]
    nums = [v for v in vals if isinstance(v, (int, float))]
    if not nums:
        return False
    return any((isinstance(v, float) and v != int(v)) or v > 12 for v in nums)

def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip() if s is not None else ""

re_section = re.compile(r"^\s*(\d+)\.\s")           # "1. ", "2. ", "3. "
re_subA    = re.compile(r"^\s*(\d+\.\d+)\s*$")        # col A "2.1"
re_subB    = re.compile(r"^\s*(\d+\.\d+)\s+\S")       # col B "1.1 ..."
re_measure = re.compile(r"^\s*(\d+\.\d+\.\d+)\s+\S")  # col B "2.1.1 ..."

items = []          # flat list of all nodes
sec = sub = mea = None
sec_n = 0

maxrow = ws.max_row
r = 4
while r <= maxrow:
    a = norm(cell(r, 1))
    b = norm(cell(r, 2))
    t = cell(r, 20)   # ผู้รับผิดชอบ
    u = cell(r, 21)   # งบประมาณ
    e = cell(r, 5)    # ตัวชี้วัด
    obj = cell(r, 3)  # วัตถุประสงค์

    # ---- Section ----
    m = re_section.match(a)
    if m and not re_subA.match(a):
        sec_n += 1
        sec = {"type":"section","id":f"S{sec_n}","code":m.group(1)+".",
               "name":a,"children":[]}
        if norm(cell(r+1,1))=="" and norm(cell(r+1,2))=="" and is_pct_row(r+1):
            sec["plan"]=monthly(r+1)
        items.append(sec); sub=mea=None; r+=1; continue

    # ---- Measure (x.y.z) in col B ----
    mm = re_measure.match(b)
    if mm and sec is not None:
        code = mm.group(1)
        pr = monthly(r) if is_pct_row(r) else (monthly(r+1) if is_pct_row(r+1) else None)
        mea = {"type":"measure","id":"M_"+code,"code":code,
               "name":re.sub(r"^\s*\d+\.\d+\.\d+\s+","",b),
               "resp":norm(t) or None,"budget":(u if isinstance(u,(int,float)) else None),
               "plan_pct":pr,"tasks":[]}
        (sub["children"] if sub else sec["children"]).append(mea)
        r+=1; continue

    # ---- Sub-plan: col A "x.y"  OR col B "x.y ..." ----
    msA = re_subA.match(a)
    msB = re_subB.match(b)
    if (msA or msB) and sec is not None:
        code = (msA.group(1) if msA else msB.group(1))
        name = re.sub(r"^\s*\d+\.\d+\s+","",b) if msB else norm(obj) or norm(e) or code
        pr = monthly(r) if is_pct_row(r) else (monthly(r+1) if is_pct_row(r+1) else None)
        sub = {"type":"subplan","id":"P_"+code,"code":code,"name":name,
               "resp":norm(t) or None,"budget":(u if isinstance(u,(int,float)) else None),
               "indicator":norm(e) or None,"objective":norm(obj) or None,
               "plan_pct":pr,"children":[],"tasks":[]}
        sec["children"].append(sub); mea=None
        r+=1; continue

    # ---- Task (plain text in col B) ----
    if b and not re.match(r"^\s*\d+\.", b):
        # skip pure number/marker rows handled implicitly; this is a named task
        sched = []
        vals = monthly(r)
        for i,v in enumerate(vals):
            if v is not None and v != 0:
                sched.append(i)   # month index task is active
        task = {"type":"task","name":b,"sched":sched}
        target = mea if mea else sub
        if target is not None:
            target["tasks"].append(task)
        r+=1; continue

    r+=1

# attach plan_pct fallback for subplans/measures with none: derive from tasks union schedule
def finalize(node):
    pass

def _cf(arr,m):
    v=0
    for i in range(m+1):
        if arr[i] is not None: v=arr[i]
    return v
for _s in items:
    if "plan" not in _s:
        _subs=[c for c in _s["children"] if c["type"]=="subplan"]
        if _subs:
            _s["plan"]=[min(100,round(sum(_cf(c.get("plan_pct") or [None]*12,m) for c in _subs)/len(_subs))) for m in range(12)]
        else:
            _s["plan"]=[None]*12
out = {"year":"2569","months":MONTHS,"sections":items}
with open(os.path.join(HERE,"plan_data.json"),"w",encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)

# ---- summary ----
def count(items):
    sp=me=ta=0
    for s in items:
        for sub in s["children"]:
            if sub["type"]=="subplan":
                sp+=1
                ta+=len(sub.get("tasks",[]))
                for mm in sub["children"]:
                    me+=1; ta+=len(mm.get("tasks",[]))
            elif sub["type"]=="measure":
                me+=1; ta+=len(sub.get("tasks",[]))
    return sp,me,ta
sp,me,ta = count(items)
print(f"Sections: {len(items)}")
for s in items:
    print(f"  {s['code']} {s['name'][:50]}  (children={len(s['children'])})")
print(f"Sub-plans: {sp}  Measures: {me}  Tasks: {ta}")
