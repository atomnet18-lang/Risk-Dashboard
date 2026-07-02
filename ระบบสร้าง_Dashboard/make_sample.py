# -*- coding: utf-8 -*-
"""สร้าง items_master(126 leaf) + MS_List_template + ข้อมูลตัวอย่าง (../data/rm_actual.csv) ระดับกิจกรรมย่อย"""
import os, csv, random
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as col
from common_items import build,leaves,CANON,DONE,DOING,NOT
HERE=os.path.dirname(os.path.abspath(__file__)); ROOT=os.path.dirname(HERE)
FORMDIR=os.path.join(ROOT,"ฟอร์ม_SharePoint"); os.makedirs(FORMDIR,exist_ok=True)

M,acts,secs,SUBPLANS=build(); lv=leaves(acts)

def ramp(sched,plan):
    """เป้าหมายสะสมรายเดือนของ leaf: ถ้ามี sched ใช้ ramp จากเดือนทำงาน; ไม่งั้นใช้ plan ของกิจกรรม"""
    if sched:
        tot=len(sched); out=[]
        for m in range(12):
            cum=sum(1 for x in sched if x<=m)
            out.append(round(100*cum/tot) if cum>0 else None)
        return out
    return [(v if v is not None else None) for v in (plan or [None]*12)]

# leaf -> (sched, plan) lookup
leafinfo={}
for a in acts:
    if a["subs"]:
        for s in a["subs"]:
            sched=[i for i in range(12) if s["bits"][i]=="1"]
            leafinfo[s["id"]]=(sched,None)
    else:
        leafinfo[a["id"]]=([],a["plan"])

# ---- items_master.csv (126 leaf อ้างอิง) ----
with open(os.path.join(HERE,"items_master.csv"),"w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f); w.writerow(["รหัสกิจกรรม","แผนหลัก","แผนงานย่อย","กิจกรรมหลัก","ชื่อกิจกรรมย่อย","ผู้รับผิดชอบ"])
    for l in lv: w.writerow([l["id"],l["sec"],l["subplan"],f'{l["parent_code"]} {l["parent_name"]}',l["name"],l["resp"]])

# ---- actuals_template.csv (หัวเปล่า) ----
with open(os.path.join(HERE,"actuals_template.csv"),"w",encoding="utf-8-sig",newline="") as f:
    csv.writer(f).writerow(CANON)

# ---- ตัวอย่างผลจริง ระดับ leaf ถึง มิ.ย. ----
random.seed(7); CUR=5; sample=[]
for l in lv:
    sched,plan=leafinfo[l["id"]]; tgt=ramp(sched,plan); last=0
    rep=l["resp"].split("/")[-1].strip() or "-"
    for mi in range(CUR+1):
        t=None
        for j in range(mi+1):
            if tgt[j] is not None: t=tgt[j]
        if t is None: continue
        act=max(last,min(100,round(t+random.choice([0,0,5,-6,-15,4,-3,-20])))); last=act
        st=DONE if act>=100 else (NOT if act==0 else DOING)
        note=f'{l["name"][:22]} — {M[mi]} ({act}%)'
        issue="ติดขั้นตอนอนุมัติ/ประสานงาน" if (t-act)>=15 else ""
        sample.append([l["id"],l["sec"],l["subplan"],f'{l["parent_code"]} {l["parent_name"]}',l["name"],l["resp"],
                       M[mi],act,st,note,issue,rep,"2026-{:02d}-25".format(mi+1)])
with open(os.path.join(ROOT,"data","rm_actual.csv"),"w",encoding="utf-8-sig",newline="") as f:
    wr=csv.writer(f); wr.writerow(CANON); wr.writerows(sample)
with open(os.path.join(HERE,"actuals_sample.csv"),"w",encoding="utf-8-sig",newline="") as f:
    wr=csv.writer(f); wr.writerow(CANON); wr.writerows(sample)

# ---- MS_List_template.xlsx ----
NAVY="16275C"; HF=Font(color="FFFFFF",bold=True,size=11); HEADF=PatternFill("solid",fgColor=NAVY)
thin=Side(style="thin",color="BBBBBB"); B=Border(thin,thin,thin,thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True); wrap=Alignment(vertical="center",wrap_text=True)
wb=openpyxl.Workbook(); ws=wb.active; ws.title="กิจกรรมย่อย_Master"
h1=["รหัสกิจกรรม","แผนหลัก","แผนงานย่อย","กิจกรรมหลัก","ชื่อกิจกรรมย่อย","ผู้รับผิดชอบ"]
ws.append(h1)
for c in range(1,len(h1)+1):
    x=ws.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
for l in lv: ws.append([l["id"],l["sec"],l["subplan"],f'{l["parent_code"]} {l["parent_name"]}',l["name"],l["resp"]])
for i,wd in enumerate([15,8,10,40,46,20],1): ws.column_dimensions[col(i)].width=wd
for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
    for c in row: c.border=B; c.alignment=wrap if c.column in(4,5) else ctr
ws.freeze_panes="A2"

ws2=wb.create_sheet("ฟอร์มกรอกผล_CANON"); ws2.append(CANON)
for c in range(1,len(CANON)+1):
    x=ws2.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
ex=lv[0]
ws2.append([ex["id"],ex["sec"],ex["subplan"],f'{ex["parent_code"]} {ex["parent_name"]}',ex["name"],ex["resp"],
            "มิ.ย.",60,DOING,"ออกแบบโครงสร้างคณะทำงานเสร็จ รออนุมัติ","","แบม","2026-06-25"])
for i,wd in enumerate([15,8,10,34,40,18,8,11,18,30,22,12,14],1): ws2.column_dimensions[col(i)].width=wd
dvm=DataValidation(type="list",formula1='"'+",".join(M)+'"',allow_blank=True)
dvp=DataValidation(type="whole",operator="between",formula1=0,formula2=100,allow_blank=True)
dvs=DataValidation(type="list",formula1=f'"{DONE},{DOING},{NOT}"',allow_blank=True)
ws2.add_data_validation(dvm);ws2.add_data_validation(dvp);ws2.add_data_validation(dvs)
dvm.add("G2:G5000");dvp.add("H2:H5000");dvs.add("I2:I5000")
for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row):
    for c in row: c.border=B; c.alignment=wrap if c.column in(4,5,10,11) else ctr
ws2.freeze_panes="A2"

ws3=wb.create_sheet("คำอธิบาย")
desc=[["คอลัมน์","ชนิด (SharePoint)","คำอธิบาย"],
 ["รหัสกิจกรรม","Text","รหัสกิจกรรมย่อย เช่น P_1.1.1, M_2.1.1.1 — เชื่อม Dashboard (ห้ามชื่อ ID/ห้ามแก้)"],
 ["แผนหลัก","Text","1. / 2. / 3. (เติมให้แล้ว)"],
 ["แผนงานย่อย","Text","เช่น 1.1, 2.4 (เติมให้แล้ว)"],
 ["กิจกรรมหลัก","Text","กิจกรรม/มาตรการแม่ของกิจกรรมย่อยนี้ (เติมให้แล้ว)"],
 ["ชื่อกิจกรรมย่อย","Text","ชื่อกิจกรรมย่อย (เติมให้แล้ว)"],
 ["ผู้รับผิดชอบ","Text/Person","เติมให้แล้ว"],
 ["เดือน","Choice","ม.ค.–ธ.ค. (กรอก)"],
 ["percent_จริง","Number 0–100","% สะสมของกิจกรรมย่อย (กรอก)"],
 ["สถานะงาน","Choice",f"{DONE}/{DOING}/{NOT} (กรอก)"],
 ["รายละเอียดการดำเนินการ","Multiple lines","สิ่งที่ทำเดือนนั้น (กรอก)"],
 ["ปัญหาอุปสรรค","Multiple lines","ปัญหา/แนวทาง (กรอก ถ้ามี)"],
 ["ผู้รายงาน","Person/Text","ผู้กรอก (กรอก)"],
 ["วันที่รายงาน","Date","วันที่กรอก (กรอก)"],
 ["—","—","Dashboard รวม % ของกิจกรรมย่อย → เป็นผลงานของกิจกรรมหลัก → แผนหลัก อัตโนมัติ"]]
for r in desc: ws3.append(r)
for c in range(1,4):
    x=ws3.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
for i,wd in enumerate([22,24,76],1): ws3.column_dimensions[col(i)].width=wd
for row in ws3.iter_rows(min_row=2,max_row=ws3.max_row):
    for c in row: c.border=B; c.alignment=wrap
wb.save(os.path.join(FORMDIR,"MS_List_template.xlsx"))
print("leaves:",len(lv),"| sample rows:",len(sample),"| CANON cols:",len(CANON))
