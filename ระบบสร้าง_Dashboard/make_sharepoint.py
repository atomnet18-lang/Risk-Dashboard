# -*- coding: utf-8 -*-
"""ไฟล์ Excel ชีตเดียว พร้อม import เข้า SharePoint List — มีครบ 48 กิจกรรม
แต่ละแถว = 1 กิจกรรม (เติมชื่อ/แผน/ผู้รับผิดชอบให้แล้ว) ช่องกรอกผลเว้นว่างให้กรอกรายเดือน"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as col
import os
from common_items import load,CANON,DONE,DOING,NOT
HERE=os.path.dirname(os.path.abspath(__file__)); ROOT=os.path.dirname(HERE)
FORMDIR=os.path.join(ROOT,"ฟอร์ม_SharePoint"); os.makedirs(FORMDIR,exist_ok=True)

M,items=load()
NAVY="16275C"; HF=Font(color="FFFFFF",bold=True,size=11); HEADF=PatternFill("solid",fgColor=NAVY)
REFILL=PatternFill("solid",fgColor="EEF2F7")  # คอลัมน์อ้างอิง (เติมให้แล้ว) สีจาง
thin=Side(style="thin",color="C7CED6"); B=Border(thin,thin,thin,thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True); wrap=Alignment(vertical="center",wrap_text=True)

wb=openpyxl.Workbook(); ws=wb.active; ws.title="ผลการดำเนินงาน"
ws.append(CANON)
for it in items:
    ws.append([it["id"],it["sec"],it["sub"],it["name"],it["resp"],"","","","","","",""])
# header style
for c in range(1,len(CANON)+1):
    x=ws.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
# cell style
for r in range(2,ws.max_row+1):
    for c in range(1,len(CANON)+1):
        cell=ws.cell(r,c); cell.border=B
        cell.alignment=wrap if c in (4,9,10) else ctr
        if c<=5: cell.fill=REFILL          # คอลัมน์อ้างอิง = พื้นจาง
        if c==12: cell.number_format="yyyy-mm-dd"
widths=[14,8,10,46,22,9,12,20,34,26,16,15]
for i,w in enumerate(widths,1): ws.column_dimensions[col(i)].width=w
ws.row_dimensions[1].height=30; ws.freeze_panes="A2"
# Excel Table จริง (ให้ SharePoint ตรวจจับ)
ref=f"A1:{col(len(CANON))}{ws.max_row}"
tbl=Table(displayName="RMActual",ref=ref)
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
ws.add_table(tbl)
# dropdowns ช่วยตอนกรอก
dvm=DataValidation(type="list",formula1='"'+",".join(M)+'"',allow_blank=True)
dvp=DataValidation(type="whole",operator="between",formula1=0,formula2=100,allow_blank=True)
dvs=DataValidation(type="list",formula1=f'"{DONE},{DOING},{NOT}"',allow_blank=True)
ws.add_data_validation(dvm);ws.add_data_validation(dvp);ws.add_data_validation(dvs)
dvm.add("F2:F5000");dvp.add("G2:G5000");dvs.add("H2:H5000")

wb.save(os.path.join(FORMDIR,"SharePoint_List_RM_2569.xlsx"))
print(f"saved SharePoint_List_RM_2569.xlsx | 1 sheet | table=RMActual | rows={len(items)} | cols={len(CANON)}")
