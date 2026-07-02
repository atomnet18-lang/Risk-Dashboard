# -*- coding: utf-8 -*-
"""แปลง บันทึกผลการดำเนินงาน_RM_2569.xlsx (เป้าหมาย+ผลจริง) -> data/rm_actual.csv + rebuild index.html"""
import os, csv, sys, subprocess
import openpyxl
from common_items import CANON, DONE, DOING, NOT
HERE=os.path.dirname(os.path.abspath(__file__)); ROOT=os.path.dirname(HERE)
SRC=os.path.join(ROOT,"บันทึกผลการดำเนินงาน_RM_2569.xlsx")
OUT=os.path.join(ROOT,"data","rm_actual.csv")
MONTHS=["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.","ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]
if not os.path.exists(SRC): sys.exit("ไม่พบไฟล์: "+SRC)
wb=openpyxl.load_workbook(SRC,data_only=True); ws=wb["บันทึกผล"]
hdr=[(ws.cell(1,c).value or "") for c in range(1,ws.max_column+1)]
H={h:i+1 for i,h in enumerate(hdr)}
tcol={m:H.get(f"เป้า {m}") for m in MONTHS}
acol={m:H.get(f"ผล {m}") for m in MONTHS}
def gi(name): return H.get(name)
def num(v):
    if v is None or v=="" : return None
    try:
        f=float(v)
        if 0<f<=1 and f!=int(f): f*=100
        return max(0,min(100,round(f)))
    except: return None

rows=[]; n=0
for r in range(2,ws.max_row+1):
    rid=ws.cell(r,gi("รหัสกิจกรรม")).value
    if not rid: continue
    base=[ws.cell(r,gi(k)).value or "" for k in ["รหัสกิจกรรม","แผนหลัก","แผนงานย่อย","กิจกรรมหลัก","ชื่อกิจกรรมย่อย","ผู้รับผิดชอบ"]]
    detail=ws.cell(r,gi("รายละเอียดการดำเนินการ (ล่าสุด)")).value or ""
    issue=ws.cell(r,gi("ปัญหาอุปสรรค")).value or ""
    reporter=ws.cell(r,gi("ผู้รายงาน")).value or (base[5].split("/")[-1].strip() if base[5] else "")
    per=[]
    for mi,m in enumerate(MONTHS):
        t=num(ws.cell(r,tcol[m]).value) if tcol[m] else None
        v=num(ws.cell(r,acol[m]).value) if acol[m] else None
        if t is not None or v is not None: per.append((mi,t,v))
    if not per: continue
    n+=1
    last_act=max([mi for mi,t,v in per if v is not None],default=-1)
    for mi,t,v in per:
        st=(DONE if v>=100 else NOT if v==0 else DOING) if v is not None else ""
        rows.append(base+[MONTHS[mi], ("" if t is None else t), ("" if v is None else v), st,
                          (detail if mi==last_act else ""),(issue if mi==last_act else ""),
                          reporter,"2026-{:02d}-25".format(mi+1)])
os.makedirs(os.path.dirname(OUT),exist_ok=True)
with open(OUT,"w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f); w.writerow(CANON); w.writerows(rows)
print(f"✓ เขียน {OUT}  | กิจกรรมย่อยมีข้อมูล {n} | แถว {len(rows)}")
# ----- สรุปรายเดือน ระดับกิจกรรม -----
SUMOUT=os.path.join(ROOT,"data","rm_summary.csv"); srows=[]
if "สรุปรายเดือน_กิจกรรม" in wb.sheetnames:
    wss=wb["สรุปรายเดือน_กิจกรรม"]
    sh=[(wss.cell(1,c).value or "") for c in range(1,wss.max_column+1)]
    SH={h:i+1 for i,h in enumerate(sh)}; scol={m:SH.get(f"สรุป {m}") for m in MONTHS}; iid=SH.get("รหัสกิจกรรม")
    for r in range(2,wss.max_row+1):
        rid=wss.cell(r,iid).value if iid else None
        if not rid: continue
        for m in MONTHS:
            cc=scol.get(m)
            if not cc: continue
            t=wss.cell(r,cc).value
            if t is not None and str(t).strip()!="": srows.append([rid,m,str(t).strip()])
with open(SUMOUT,"w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f); w.writerow(["รหัสกิจกรรม","เดือน","สรุปผล"]); w.writerows(srows)
print(f"✓ เขียน {SUMOUT} | สรุป {len(srows)} รายการ")
# ----- รายงานประจำเดือน -----
MONOUT=os.path.join(ROOT,"data","rm_monthly.csv"); mrows=[]
if "รายงานประจำเดือน" in wb.sheetnames:
    wsm=wb["รายงานประจำเดือน"]
    for r in range(2,wsm.max_row+1):
        mm=wsm.cell(r,1).value; tt=wsm.cell(r,2).value
        if mm and tt is not None and str(tt).strip()!="": mrows.append([str(mm).strip(),str(tt).strip()])
with open(MONOUT,"w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f); w.writerow(["เดือน","รายงาน"]); w.writerows(mrows)
print(f"✓ เขียน {MONOUT} | รายงานเดือน {len(mrows)} รายการ")
print("rebuild index.html ...")
subprocess.run([sys.executable, os.path.join(HERE,"build_dashboard.py")], check=True)
print("เสร็จ — git add -A && git commit && git push")
