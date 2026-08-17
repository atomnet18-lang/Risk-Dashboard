# -*- coding: utf-8 -*-
"""เทมเพลตกรอกผล: 1 แถว = 1 กิจกรรมย่อย · มีคอลัมน์ "เป้าหมาย" (แก้เองได้) + "ผลจริง" รายเดือน
ออก -> ../บันทึกผลการดำเนินงาน_RM_2569.xlsx"""
import os, random
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as col
from common_items import build,leaves
HERE=os.path.dirname(os.path.abspath(__file__)); ROOT=os.path.dirname(HERE)
M,acts,secs,SUBPLANS=build(); lv=leaves(acts)
info={}
for a in acts:
    if a["subs"]:
        for s in a["subs"]:
            info[s["id"]]=("bits",[i for i in range(12) if s["bits"][i]=="1"])
    else: info[a["id"]]=("plan",a["plan"])
def target_cum(idd):
    kind,d=info[idd]
    if kind=="bits":
        if not d: return [None]*12
        raw=[ (round(100*sum(1 for x in d if x<=m)/len(d)) if any(x<=m for x in d) else None) for m in range(12)]
    else:
        raw=[(v if v is not None else None) for v in (d or [None]*12)]
    out=[]; last=None
    for v in raw:
        if v is not None: last=v
        out.append(last)   # carry-forward (None ก่อนเริ่ม)
    return out

NAVY="16275C"; HF=Font(color="FFFFFF",bold=True,size=10); HEADF=PatternFill("solid",fgColor=NAVY)
TGTH=PatternFill("solid",fgColor="8a6d1f"); ACTH=PatternFill("solid",fgColor="1e3a8a")
REF=PatternFill("solid",fgColor="EEF2F7"); TGF=PatternFill("solid",fgColor="FBF3DD"); ACF=PatternFill("solid",fgColor="EAF1FB")
thin=Side(style="thin",color="C7CED6"); B=Border(thin,thin,thin,thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True); wrap=Alignment(vertical="center",wrap_text=True)
title=Font(bold=True,size=13,color=NAVY)

WEIGHT={a["id"]:a["weight"] for a in acts}
wb=openpyxl.Workbook(); ws=wb.active; ws.title="บันทึกผล"
HEAD=["รหัสกิจกรรม","แผนหลัก","แผนงานย่อย","รหัส","ชื่อกิจกรรม","ผู้รับผิดชอบ"]\
     +[f"เป้า {m}" for m in M]+[f"ผล {m}" for m in M]+["รายละเอียดการดำเนินการ (ล่าสุด)","ปัญหาอุปสรรค","ผู้รายงาน","ถ่วงน้ำหนัก(%)"]
ws.append(HEAD)
random.seed(7); CUR=5
for l in lv:
    tg=target_cum(l["id"])
    tcells=[("" if v is None else v) for v in tg]
    # prefill actual (ตัวอย่างถึง มิ.ย.)
    acells=[""]*12; last=0; latest=-1
    for mi in range(CUR+1):
        t=tg[mi]
        if t is None: continue
        act=max(last,min(100,round(t+random.choice([0,0,5,-6,-15,4,-3,-20])))); last=act; latest=mi
        acells[mi]=act
    detail=f'ดำเนินการ {l["name"][:20]} (ถึง {M[latest]})' if latest>=0 else ""
    issue="ล่าช้ากว่าแผน ติดขั้นตอนอนุมัติ" if (latest>=0 and tg[latest] is not None and (tg[latest]-acells[latest])>=15) else ""
    rep=l["resp"].split("/")[-1].strip() or "-"
    ws.append([l["id"],l["sec"],l["subplan"],l["parent_code"],l["name"],l["resp"]]+tcells+acells+[detail,issue,rep,WEIGHT.get(l["id"],0)])
# styling
NC=len(HEAD)
WCOL=NC  # คอลัมน์ "ถ่วงน้ำหนัก(%)" (อ้างอิงเท่านั้น ใช้ในสูตร SUMPRODUCT ของชีตสรุป)
for c in range(1,NC+1):
    x=ws.cell(1,c); x.font=HF; x.alignment=ctr; x.border=B
    x.fill=TGTH if 7<=c<=18 else ACTH if 19<=c<=30 else HEADF
for r in range(2,ws.max_row+1):
    for c in range(1,NC+1):
        cell=ws.cell(r,c); cell.border=B; cell.alignment=wrap if c in (4,5,31,32) else ctr
        if c==WCOL: cell.fill=REF; cell.number_format="0.00"
        elif c<=6: cell.fill=REF
        elif 7<=c<=18: cell.fill=TGF
        elif 19<=c<=30: cell.fill=ACF
widths=[15,7,9,28,36,16]+[7]*24+[32,24,13,14]
for i,w in enumerate(widths,1): ws.column_dimensions[col(i)].width=w
ws.row_dimensions[1].height=30; ws.freeze_panes="G2"
tbl=Table(displayName="RMEntry",ref=f"A1:{col(NC)}{ws.max_row}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=False); ws.add_table(tbl)
dvp=DataValidation(type="whole",operator="between",formula1=0,formula2=100,allow_blank=True)
ws.add_data_validation(dvp); dvp.add(f"G2:AD{ws.max_row}")
WCOL_LETTER=col(WCOL); LASTROW=ws.max_row; PLANCOL="B"; TGT_FIRST_COL=col(7); TGT_LAST_COL=col(18)

# Sheet 2: วิธีใช้
w3=wb.create_sheet("วิธีใช้")
steps=[["วิธีกรอกและนำเข้า Dashboard",""],["",""],
 ["1)","ชีต \"บันทึกผล\" — แต่ละแถวคือกิจกรรม 1 ข้อ (48 แถว) (คอลัมน์ 1-6 และ \"ถ่วงน้ำหนัก(%)\" เติมให้แล้ว ห้ามแก้รหัส/น้ำหนัก)"],
 ["2)","กลุ่ม \"เป้า ม.ค.–ธ.ค.\" (พื้นเหลือง) = ค่าเป้าหมาย % สะสม — พรีฟิลตามแผนให้แล้ว แก้ได้ตามต้องการ"],
 ["3)","กลุ่ม \"ผล ม.ค.–ธ.ค.\" (พื้นฟ้า) = % ผลจริงสะสม กรอกเฉพาะเดือนที่อัปเดต"],
 ["","   เดือนที่เว้นว่าง ระบบถือว่าคงค่าล่าสุด (ทั้งเป้าหมายและผลจริง)"],
 ["4)","รายละเอียด/ปัญหา/ผู้รายงาน = ของเดือนล่าสุด"],
 ["",""],
 ["แปลงเข้า GitHub:",""],
 ["5)","python3 ระบบสร้าง_Dashboard/convert_to_csv.py   → สร้าง data/rm_actual.csv + อัปเดต index.html"],
 ["6)","git add -A && git commit -m \"update\" && git push"],
 ["",""],
 ["หมายเหตุ","เป้าหมายทุกระดับใน Dashboard (48 กิจกรรม, 3 กล่องแผนหลัก, กราฟ S-Curve/แท่ง) อ้างอิงจากคอลัมน์ \"เป้า\" ในไฟล์นี้ทั้งหมด"],
 ["","ตัวเลขเป้าหมายระดับแผนหลักคือค่าเฉลี่ยถ่วงน้ำหนัก (ตามคอลัมน์ \"ถ่วงน้ำหนัก(%)\") ของกิจกรรมในแผนนั้น — ดูสูตรจริงได้ที่ชีต \"สรุปเป้าหมายตามแผน\""],
 ["","ชีต \"สรุปเป้าหมายตามแผน\" คำนวณสดด้วยสูตร SUMPRODUCT จากคอลัมน์เป้า+น้ำหนัก แก้เป้าหมายกิจกรรมใดก็ตาม ตัวเลขแผนหลักในชีตนี้จะขยับให้อัตโนมัติ (ตรงกับที่ Dashboard จะแสดง)"],
 ["","สถานะงานคำนวณจากผลจริง: 100=เสร็จสิ้น · 1-99=อยู่ระหว่างดำเนินการ · 0/ว่าง=ยังไม่ดำเนินการ"]]
for r in steps: w3.append(r)
w3["A1"].font=title
for r in range(1,w3.max_row+1):
    w3.cell(r,1).alignment=Alignment(vertical="top",horizontal="right")
    w3.cell(r,2).alignment=Alignment(vertical="top",wrap_text=True)
w3["A9"].font=title
w3.column_dimensions["A"].width=10; w3.column_dimensions["B"].width=92

# ===== Sheet: สรุปรายเดือน (ระดับกิจกรรม 48 หัวข้อ) =====
def _cf(arr,m):
    v=0
    for i in range(m+1):
        if arr[i] is not None: v=arr[i]
    return v
wss=wb.create_sheet("สรุปรายเดือน_กิจกรรม")
SH=["รหัสกิจกรรม","รหัส","ชื่อกิจกรรม/มาตรการ","แผนหลัก"]+[f"สรุป {m}" for m in M]
wss.append(SH)
for a in acts:
    cells=[""]*12
    for mi in range(6):
        if _cf(a["plan"],mi)>0:
            cells[mi]=f'ดำเนินการ {a["name"][:30]} ตามแผนในเดือน{M[mi]} (สรุปย่อ — แก้เป็นเนื้อหาจริงได้)'
    wss.append([a["id"],a["code"],a["name"],a["section"]]+cells)
for c in range(1,len(SH)+1):
    x=wss.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
for r in range(2,wss.max_row+1):
    for c in range(1,len(SH)+1):
        cell=wss.cell(r,c); cell.border=B; cell.alignment=wrap
        if c<=4: cell.fill=REF
for i,w in enumerate([15,8,42,8]+[40]*12,1): wss.column_dimensions[col(i)].width=w
wss.freeze_panes="E2"; wss.row_dimensions[1].height=22

# ===== Sheet: รายงานประจำเดือน (ภาพรวมทั้งเดือน 1 ข้อความต่อเดือน) =====
wsm=wb.create_sheet("รายงานประจำเดือน")
wsm.append(["เดือน","รายงานภาพรวมประจำเดือน"])
for mi,m in enumerate(M):
    txt=f"ภาพรวมการดำเนินงานตามแผนบริหารความเสี่ยงประจำเดือน{m} — (แก้เป็นเนื้อหาสรุปจริง)" if mi<6 else ""
    wsm.append([m,txt])
for c in range(1,3):
    x=wsm.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
for r in range(2,wsm.max_row+1):
    for c in range(1,3):
        cell=wsm.cell(r,c); cell.border=B
        cell.alignment=ctr if c==1 else wrap
        if c==1: cell.fill=REF
wsm.column_dimensions["A"].width=10; wsm.column_dimensions["B"].width=95; wsm.row_dimensions[1].height=22

# ===== Sheet: สรุปเป้าหมายตามแผน (สูตรผูกอัตโนมัติจากชีต "บันทึกผล") =====
# เป้าหมายสะสมถ่วงน้ำหนักรายเดือน ของแต่ละแผนหลัก + รวมทั้งหมด
# คำนวณสด (SUMPRODUCT) จากคอลัมน์ "เป้า <เดือน>" x "ถ่วงน้ำหนัก(%)" ในชีต "บันทึกผล"
# ตัวเลขนี้คือค่าที่ Dashboard จะแสดงในกล่องแผนหลัก 3 กล่อง + กราฟ S-Curve/แท่ง (เส้น/แท่งเป้าหมาย)
wst=wb.create_sheet("สรุปเป้าหมายตามแผน")
TH=["แผนหลัก","รหัสแผน"]+M
wst.append(TH)
def _plan_target_formula(code_cell):
    parts=[]
    for mc in range(7,19):
        L=col(mc)
        f=(f'=IFERROR(SUMPRODUCT((บันทึกผล!${PLANCOL}$2:${PLANCOL}${LASTROW}={code_cell})*'
           f'บันทึกผล!${WCOL_LETTER}$2:${WCOL_LETTER}${LASTROW}*N(บันทึกผล!{L}$2:{L}${LASTROW}))/'
           f'SUMPRODUCT((บันทึกผล!${PLANCOL}$2:${PLANCOL}${LASTROW}={code_cell})*'
           f'บันทึกผล!${WCOL_LETTER}$2:${WCOL_LETTER}${LASTROW}),"")')
        parts.append(f)
    return parts
def _overall_target_formula():
    parts=[]
    for mc in range(7,19):
        L=col(mc)
        f=(f'=IFERROR(SUMPRODUCT(บันทึกผล!${WCOL_LETTER}$2:${WCOL_LETTER}${LASTROW}*'
           f'N(บันทึกผล!{L}$2:{L}${LASTROW}))/SUMPRODUCT(บันทึกผล!${WCOL_LETTER}$2:${WCOL_LETTER}${LASTROW}),"")')
        parts.append(f)
    return parts
for s in secs:
    r=wst.max_row+1
    wst.append([s["name"],s["code"]]+_plan_target_formula(f"$B{r}"))
r=wst.max_row+1
wst.append(["รวมทั้งหมด (ถ่วงน้ำหนักทุกกิจกรรม)","*"]+_overall_target_formula())
for c in range(1,len(TH)+1):
    x=wst.cell(1,c); x.fill=HEADF; x.font=HF; x.alignment=ctr; x.border=B
for r in range(2,wst.max_row+1):
    for c in range(1,len(TH)+1):
        cell=wst.cell(r,c); cell.border=B
        cell.alignment=wrap if c==1 else ctr
        if c<=2: cell.fill=REF
        else: cell.fill=TGF; cell.number_format="0.0"
wst.column_dimensions["A"].width=34; wst.column_dimensions["B"].width=9
for i in range(3,len(TH)+1): wst.column_dimensions[col(i)].width=9
wst.row_dimensions[1].height=22
wst.freeze_panes="C2"

out=os.path.join(ROOT,"บันทึกผลการดำเนินงาน_RM_2569.xlsx")
wb.save(out)
print("saved:",os.path.basename(out),"| cols=",NC,"| rows=",len(lv))
